#!/usr/bin/env python3
"""
build_bom.py — Assemble the Blazepod clone Bill of Materials (.xlsx).

WHAT IT DOES
    Joins the static clone-component map (which subsystem, which suggested part)
    with the live Iranian pricing pulled by iran_parts.py, and renders a clean
    spreadsheet with one row per component.

USAGE
    python tools/build_bom.py --parts .tmp/iran_parts.json --out export/blazepod-clone-bom.xlsx

INPUT (.tmp/iran_parts.json)
    Output of tools/iran_parts.py — list of {id, query, results, needs_confirmation}.
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "Missing dependency. Install with:\n  pip install -r requirements.txt\n"
    )
    raise

# ---------------------------------------------------------------------------
# Static component map — the engineering choice per subsystem.
# Keys correspond to the `id` values in .tmp/part_queries.json.
# ---------------------------------------------------------------------------
COMPONENT_MAP = {
    "mcu": {
        "subsystem": "Processor (Microcontroller)",
        "role": "Runs the pod firmware: receives BLE commands, drives the LEDs, "
                "reads the tap sensor, times reactions.",
        "blazepod_spec": "Custom BLE SoC (likely nRF52 / similar; see FCC ID "
                         "2AQTQBLAZEPOD internal photos).",
        "suggested_part": "ESP32-C3 (RISC-V, built-in BLE 5) — cheap, well-"
                          "documented, runs WLED or custom firmware.",
    },
    "leds": {
        "subsystem": "LEDs (the light)",
        "role": "Produces the visible colored light the athlete taps.",
        "blazepod_spec": "Full-color RGB, bright, even diffused glow across the "
                         "pod face. Exact LED array not public.",
        "suggested_part": "WS2812B addressable RGB LED ring (or NeoPixel strip), "
                          "8–12 LEDs, driven from one GPIO pin.",
    },
    "battery": {
        "subsystem": "Battery",
        "role": "Stores energy to power the pod away from the charger.",
        "blazepod_spec": "3.7V ~600 mAh Li-Po, non-replaceable, ~8h runtime / "
                         "~200h standby.",
        "suggested_part": "3.7V 600–800 mAh Li-Po pouch cell with protection PCB.",
    },
    "charger": {
        "subsystem": "Charging circuit",
        "role": "Safely charges the Li-Po from USB and protects against "
                "overcharge/discharge.",
        "blazepod_spec": "Stackable charging base; pods charge when stacked.",
        "suggested_part": "TP4056 Li-Ion charger module (USB-C variant) with "
                          "protection. For stackable charging add pogo pins.",
    },
    "touch_sensor": {
        "subsystem": "Tap / touch sensor",
        "role": "Detects when the athlete hits the pod, with low latency.",
        "blazepod_spec": "Sensitive tap detection through the shell. Likely "
                         "capacitive or force-based.",
        "suggested_part": "TTP223 capacitive touch module (cheap, digital out) "
                          "OR FSR-402 force-sensitive resistor for impact.",
    },
    "ble": {
        "subsystem": "Bluetooth (BLE)",
        "role": "Wireless link between pod and phone app.",
        "blazepod_spec": "Bluetooth Low Energy; multiple pods paired "
                         "simultaneously; star topology to phone.",
        "suggested_part": "Integrated in ESP32-C3 (no separate module needed).",
    },
    "housing": {
        "subsystem": "Housing / shell",
        "role": "Holds the electronics, diffuses the LED light evenly, survives "
                "being hit, mounts to surfaces.",
        "blazepod_spec": "ABS / polycarbonate shell with silicone edge; "
                         "weather-resistant (IP rated); disc shape.",
        "suggested_part": "3D-printed PETG/ABS disc + silicone/TPU edge ring + "
                          "translucent diffuser over LEDs. Optional suction base.",
    },
    "diffuser": {
        "subsystem": "Light diffuser",
        "role": "Spreads the point LEDs into an even glow across the pod face.",
        "blazepod_spec": "Translucent dome giving even, soft color.",
        "suggested_part": "3D-printed / cast translucent diffuser, or frosted "
                          "acrylic disc, spaced ~5mm above LEDs.",
    },
}

COLUMNS = [
    ("Subsystem", 28),
    ("Role", 48),
    ("Blazepod spec", 40),
    ("Suggested part", 44),
    ("Iran source (shop)", 20),
    ("Price (Toman)", 16),
    ("In stock", 10),
    ("Source URL", 38),
    ("Status", 16),
]

HEADER_FILL = PatternFill("solid", fgColor="1B4E9E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
WARN_FONT = Font(color="856404", bold=True)
OK_FILL = PatternFill("solid", fgColor="D4EDDA")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _load(parts_path):
    with Path(parts_path).open(encoding="utf-8") as f:
        return json.load(f)


def _best_result(results):
    """Pick the best priced result; prefer one with a non-null price."""
    priced = [r for r in results if r.get("price_toman")]
    if priced:
        # Cheapest available
        return min(priced, key=lambda r: r["price_toman"])
    return results[0] if results else None


def build(parts_path, out_path):
    parts = _load(parts_path)
    # Index parts by id for quick lookup
    by_id = {p["id"]: p for p in parts}

    wb = Workbook()
    ws = wb.active
    ws.title = "Blazepod Clone BOM"

    # Header row ------------------------------------------------------------
    for col, (name, _w) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    # Set widths
    for col, (_name, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Body ------------------------------------------------------------------
    row = 2
    for cid, meta in COMPONENT_MAP.items():
        part = by_id.get(cid)
        best = _best_result(part["results"]) if part else None

        shop = best.get("shop", "") if best else ""
        price = best.get("price_toman") if best else None
        stock = "Yes" if (best and best.get("in_stock")) else ("No" if best else "")
        url = best.get("url", "") if best else ""

        if not best:
            status = "NOT_FOUND"
        elif price is None:
            status = "NEEDS_CONFIRMATION"
        else:
            status = "OK"

        values = [
            meta["subsystem"],
            meta["role"],
            meta["blazepod_spec"],
            meta["suggested_part"],
            shop,
            price if price is not None else "",
            stock,
            url,
            status,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if status == "NEEDS_CONFIRMATION":
                cell.fill = WARN_FILL
            elif status == "OK":
                # Light green only on status cell to keep readability
                if col == 9:
                    cell.fill = OK_FILL

        # Hyperlink the URL
        if url:
            url_cell = ws.cell(row=row, column=8)
            url_cell.hyperlink = url
            url_cell.font = Font(color="1A4E9E", underline="single")

        ws.row_dimensions[row].height = 60
        row += 1

    # Freeze header
    ws.freeze_panes = "A2"

    # Notes sheet -----------------------------------------------------------
    notes = wb.create_sheet("Notes")
    notes.column_dimensions["A"].width = 100
    lines = [
        ("Blazepod Clone — Bill of Materials", True, 14),
        (f"Generated: {datetime.now().isoformat(timespec='seconds')}", False, 10),
        ("", False, 10),
        ("Status legend:", True, 11),
        ("OK — a price was found automatically.", False, 10),
        ("NEEDS_CONFIRMATION — the shop page was reachable but a price could "
         "not be parsed. Verify manually at the source URL.", False, 10),
        ("NOT_FOUND — no result returned by any shop for this component. "
         "Broaden the search term (see part_queries.json aliases).", False, 10),
        ("", False, 10),
        ("Important:", True, 11),
        ("Prices are in Iranian Toman and reflect the moment they were fetched. "
         "Iranian electronics shops frequently rate-limit or block automated "
         "access, so some rows will require manual confirmation.", False, 10),
        ("Quantities shown are per single pod. Multiply by the number of pods "
         "in your kit (typically 4 or 6).", False, 10),
        ("", False, 10),
        ("Next steps:", True, 11),
        ("1. Confirm any NEEDS_CONFIRMATION / NOT_FOUND rows in person at the "
         "shop or by phone.", False, 10),
        ("2. Decide pod count (4 vs 6) and multiply BOM quantities.", False, 10),
        ("3. Prototype one pod before committing to a full kit.", False, 10),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        c = notes.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = WRAP

    out_path = Path(out_path).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    sys.stdout.write(f"[build_bom] Wrote {out_path}\n")
    sys.stdout.write(
        "[build_bom] Rows: "
        f"{sum(1 for _ in COMPONENT_MAP)} total, "
        f"{sum(1 for p in parts if not p.get('needs_confirmation'))} priced, "
        f"{sum(1 for p in parts if p.get('needs_confirmation'))} need confirmation.\n"
    )
    return out_path


def main():
    p = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    p.add_argument("--parts", required=True, help="Path to iran_parts.json")
    p.add_argument("--out", required=True, help="Output .xlsx path")
    args = p.parse_args()
    build(args.parts, args.out)


if __name__ == "__main__":
    main()
